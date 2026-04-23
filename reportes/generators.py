"""
reportes/generators.py
Genera archivos PDF y Excel para cada módulo del sistema.
"""
import io
from datetime import datetime

# ── Excel ──────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PDF ────────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ── Paleta Mine Inventory ──────────────────────────────────────────────────
DARK   = colors.HexColor('#1B2021')
NAVY   = colors.HexColor('#094D92')
RUST   = colors.HexColor('#98473E')
CREAM  = colors.HexColor('#EFECCA')
SAGE   = colors.HexColor('#71816D')
BGBODY = colors.HexColor('#F2F0EB')

EXCEL_DARK  = '1B2021'
EXCEL_NAVY  = '094D92'
EXCEL_RUST  = '98473E'
EXCEL_CREAM = 'EFECCA'
EXCEL_SAGE  = '71816D'
EXCEL_LIGHT = 'F2F0EB'


# ══════════════════════════════════════════════════════════════════
#  EXCEL helpers
# ══════════════════════════════════════════════════════════════════

def _excel_border():
    thin = Side(style='thin', color='D0CCC4')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _excel_workbook(modulo_label, headers, rows):
    """Crea un workbook Excel estilizado y devuelve un BytesIO."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = modulo_label[:31]

    # ── Fila de título ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f'MINE INVENTORY — {modulo_label.upper()}'
    title_cell.font      = Font(name='Calibri', bold=True, size=14, color=EXCEL_CREAM)
    title_cell.fill      = PatternFill('solid', fgColor=EXCEL_DARK)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── Fila de fecha ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    date_cell = ws.cell(row=2, column=1)
    date_cell.value = f'Generado el {datetime.now():%d/%m/%Y a las %H:%M}'
    date_cell.font      = Font(name='Calibri', italic=True, size=10, color=EXCEL_SAGE)
    date_cell.fill      = PatternFill('solid', fgColor='F9F7F3')
    date_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # ── Encabezados ──
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(name='Calibri', bold=True, size=10, color=EXCEL_CREAM)
        cell.fill      = PatternFill('solid', fgColor=EXCEL_NAVY)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _excel_border()
    ws.row_dimensions[3].height = 20

    # ── Datos ──
    for r_idx, row in enumerate(rows, start=4):
        fill_color = EXCEL_LIGHT if r_idx % 2 == 0 else 'FFFFFF'
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = Font(name='Calibri', size=10)
            cell.fill      = PatternFill('solid', fgColor=fill_color)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border    = _excel_border()
        ws.row_dimensions[r_idx].height = 16

    # ── Ajustar ancho de columnas ──
    for col_idx, header in enumerate(headers, start=1):
        col_values = [str(header)] + [str(r[col_idx - 1]) for r in rows]
        max_len = max(len(v) for v in col_values)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════
#  PDF helpers
# ══════════════════════════════════════════════════════════════════

def _pdf_doc(modulo_label, headers, rows):
    """Crea un PDF estilizado y devuelve un BytesIO."""
    buf = io.BytesIO()
    page = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        buf, pagesize=page,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        'MineTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=CREAM,
        backColor=DARK,
        alignment=TA_CENTER,
        spaceAfter=0,
        spaceBefore=0,
        leading=24,
        leftIndent=-12,
        rightIndent=-12,
        borderPadding=(8, 12, 8, 12),
    )
    style_sub = ParagraphStyle(
        'MineSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=SAGE,
        alignment=TA_CENTER,
        spaceAfter=10,
    )
    style_cell = ParagraphStyle(
        'MineCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
    )

    elements = []

    # Título
    elements.append(Paragraph(f'MINE INVENTORY — {modulo_label.upper()}', style_title))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f'Generado el {datetime.now():%d/%m/%Y a las %H:%M} &nbsp;|&nbsp; {len(rows)} registro(s)',
        style_sub,
    ))
    elements.append(HRFlowable(width='100%', thickness=2, color=RUST, spaceAfter=10))

    # Tabla
    avail_w = page[0] - 3*cm
    col_w   = avail_w / len(headers)

    table_data = [[Paragraph(f'<b>{h}</b>', ParagraphStyle(
        'H', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white,
        alignment=TA_CENTER, leading=10,
    )) for h in headers]]

    for row in rows:
        table_data.append([
            Paragraph(str(v) if v is not None else '—', style_cell)
            for v in row
        ])

    t = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
    t.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND',  (0, 0), (-1, 0), NAVY),
        ('TEXTCOLOR',   (0, 0), (-1, 0), colors.white),
        ('ALIGN',       (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 8),
        ('TOPPADDING',  (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        # Filas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, BGBODY]),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('TOPPADDING',  (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        # Bordes
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#D0CCC4')),
        ('LINEBELOW',   (0, 0), (-1, 0), 1.5, RUST),
    ]))

    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════
#  MÓDULOS — definición de headers y extracción de filas
# ══════════════════════════════════════════════════════════════════

def _data_inventario():
    from inventario.models import Producto
    qs = Producto.objects.select_related('categoria').all()
    headers = ['SKU', 'Nombre', 'Categoría', 'Stock', 'Descripción', 'Creado en']
    rows = [
        (
            p.codigo_sku,
            p.nombre,
            p.categoria.nombre if p.categoria else '—',
            p.stock,
            p.descripcion or '—',
            p.creado_en.strftime('%d/%m/%Y'),
        )
        for p in qs
    ]
    return headers, rows, 'Inventario'


def _data_prestamos():
    from prestamo.models import Prestamo
    qs = Prestamo.objects.prefetch_related('items__producto').all()
    headers = ['#', 'Usuario', 'Productos', 'Cantidades', 'Estado', 'Fecha préstamo', 'Observaciones']
    rows = []
    for p in qs:
        items = p.items.all()
        productos  = ', '.join(i.producto.nombre for i in items) or '—'
        cantidades = ', '.join(str(i.cantidad)   for i in items) or '—'
        rows.append((
            p.pk,
            p.usuario,
            productos,
            cantidades,
            p.get_estado_display(),
            p.fecha_prestamo.strftime('%d/%m/%Y %H:%M'),
            p.observaciones or '—',
        ))
    return headers, rows, 'Préstamos'


def _data_devoluciones():
    from devoluciones.models import Devolucion
    qs = Devolucion.objects.select_related('prestamo').prefetch_related('items__producto').all()
    headers = ['#', 'Préstamo', 'Usuario', 'Tipo', 'Ítems devueltos', 'Motivo', 'Estado', 'Fecha']
    rows = []
    for d in qs:
        items_str = ', '.join(
            f'{i.producto.nombre} x{i.cantidad}' for i in d.items.all()
        ) or '—'
        motivo = (d.motivo or '')
        motivo = motivo[:60] + ('…' if len(motivo) > 60 else '')
        rows.append((
            d.pk,
            f'#{d.prestamo_id}',
            d.prestamo.usuario,
            'Total' if d.devolucion_total else 'Parcial',
            items_str,
            motivo or '—',
            d.get_estado_display(),
            d.fecha_creacion.strftime('%d/%m/%Y'),
        ))
    return headers, rows, 'Devoluciones'


def _data_mantenimiento():
    from mantenimiento.models import EstadoHerramienta
    qs = EstadoHerramienta.objects.all()
    headers = ['Código', 'Herramienta', 'Descripción', 'Categoría', 'Estado']
    rows = [
        (
            h.codigo,
            h.nombre_herramienta,
            h.descripcion,
            h.get_categoria_display(),
            h.get_estado_display(),
        )
        for h in qs
    ]
    return headers, rows, 'Mantenimiento'


def _data_almacenamiento():
    from almacenamiento.models import Almacen, Estante
    almacenes = Almacen.objects.prefetch_related('estante_set').all()
    headers = ['Almacén', 'Capacidad almacén', 'Código estante', 'Capacidad estante', 'Detalles estante']
    rows = []
    for a in almacenes:
        estantes = a.estante_set.all()
        if estantes:
            for e in estantes:
                rows.append((a.nombre, a.capacidad or '—', e.codigo, e.capacidad or '—', e.detalles or '—'))
        else:
            rows.append((a.nombre, a.capacidad or '—', '—', '—', '—'))
    return headers, rows, 'Almacenamiento'


def _data_usuarios():
    from usuario.models import Usuario
    qs = Usuario.objects.select_related('id_rol').all()
    headers = ['Documento', 'Tipo', 'Nombre completo', 'Correo', 'Teléfono', 'Rol']
    rows = [
        (
            u.numero_documento,
            u.get_tipo_documento_display(),
            u.nombre_completo,
            u.correo,
            u.telefono or '—',
            u.id_rol.nombre if u.id_rol else '—',
        )
        for u in qs
    ]
    return headers, rows, 'Usuarios'


# ══════════════════════════════════════════════════════════════════
#  DISPATCHER
# ══════════════════════════════════════════════════════════════════

MODULO_MAP = {
    'inventario':     _data_inventario,
    'prestamos':      _data_prestamos,
    'devoluciones':   _data_devoluciones,
    'mantenimiento':  _data_mantenimiento,
    'almacenamiento': _data_almacenamiento,
    'usuarios':       _data_usuarios,
}


def generar_reporte(modulo, formato):
    """
    Genera el reporte y devuelve (buffer_bytes, content_type, filename, total_registros).
    """
    fn = MODULO_MAP.get(modulo)
    if not fn:
        raise ValueError(f'Módulo desconocido: {modulo}')

    headers, rows, label = fn()
    total = len(rows)
    ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
    slug  = modulo.replace(' ', '_')

    if formato == 'excel':
        buf          = _excel_workbook(label, headers, rows)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename     = f'reporte_{slug}_{ts}.xlsx'
    elif formato == 'pdf':
        buf          = _pdf_doc(label, headers, rows)
        content_type = 'application/pdf'
        filename     = f'reporte_{slug}_{ts}.pdf'
    else:
        raise ValueError(f'Formato desconocido: {formato}')

    return buf, content_type, filename, total